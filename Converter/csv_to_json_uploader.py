"""
CSV / Excel → JSON  ·  GitHub Uploader
Developed by Hood College Coding Club
Requirements: pip install requests openpyxl
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import json
import os
import re
import base64
import logging
import threading
import requests
from datetime import datetime
from pathlib import Path


# ─── File-based logger ───────────────────────────────────────────────────────
LOG_DIR  = Path.home() / "HoodCC_Logs"
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"uploader_{datetime.now().strftime('%Y%m%d')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
    ]
)
file_logger = logging.getLogger("uploader")


def log_to_file(level: str, msg: str):
    clean = msg.replace("✅", "OK").replace("❌", "ERR").replace("⚠", "WARN")
    getattr(file_logger, level.lower(), file_logger.info)(clean)


# ─── GitHub URL parser ───────────────────────────────────────────────────────
def parse_github_url(value: str):
    value = value.strip().rstrip("/")
    m = re.match(r"(?:https?://)?github\.com/([^/]+)(?:/([^/]+?))?(?:\.git)?$", value)
    if m:
        return m.group(1), (m.group(2) or "")
    m = re.match(r"([^/\s]+)/([^/\s]+)", value)
    if m:
        return m.group(1), m.group(2)
    return value, ""


# ─── Color Palette ───────────────────────────────────────────────────────────
BG        = "#0f1117"
SURFACE   = "#1a1d27"
CARD      = "#21253a"
ACCENT    = "#4f8ef7"
ACCENT2   = "#a78bfa"
SUCCESS   = "#34d399"
WARNING   = "#fbbf24"
ERROR     = "#f87171"
TEXT      = "#e2e8f0"
SUBTEXT   = "#8892a4"
BORDER    = "#2d3352"
FONT_MONO = ("Courier New", 10)
FONT_UI   = ("Segoe UI", 10)
FONT_H1   = ("Segoe UI", 15, "bold")
FONT_H2   = ("Segoe UI", 11, "bold")
FONT_SM   = ("Segoe UI", 9)
FONT_XS   = ("Segoe UI", 8)


# ─── Step tracker widget ─────────────────────────────────────────────────────
STEPS = [
    ("1", "Select File"),
    ("2", "Configure GitHub"),
    ("3", "Test Connection"),
    ("4", "Convert"),
    ("5", "Upload"),
    ("6", "Done ✓"),
]

STEP_IDLE    = "idle"
STEP_ACTIVE  = "active"
STEP_DONE    = "done"
STEP_ERROR   = "error"


class StepBar(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=SURFACE, **kw)
        self._labels = []
        self._dots   = []
        self._states = [STEP_IDLE] * len(STEPS)
        for i, (num, name) in enumerate(STEPS):
            col = tk.Frame(self, bg=SURFACE)
            col.pack(side="left", padx=6)
            dot = tk.Label(col, text=num, width=2, font=("Segoe UI", 9, "bold"),
                           bg=BORDER, fg=SUBTEXT, relief="flat")
            dot.pack()
            lbl = tk.Label(col, text=name, font=FONT_XS, bg=SURFACE, fg=SUBTEXT)
            lbl.pack()
            self._dots.append(dot)
            self._labels.append(lbl)
            if i < len(STEPS) - 1:
                tk.Label(self, text="─────", bg=SURFACE, fg=BORDER,
                         font=("Segoe UI", 9)).pack(side="left", pady=(0, 14))

    def set(self, step_index: int, state: str):
        """Set a step's visual state: idle / active / done / error"""
        if step_index >= len(STEPS):
            return
        self._states[step_index] = state
        dot = self._dots[step_index]
        lbl = self._labels[step_index]
        if state == STEP_ACTIVE:
            dot.config(bg=ACCENT,   fg="white")
            lbl.config(fg=TEXT)
        elif state == STEP_DONE:
            dot.config(bg=SUCCESS,  fg=BG)
            lbl.config(fg=SUCCESS)
        elif state == STEP_ERROR:
            dot.config(bg=ERROR,    fg=BG)
            lbl.config(fg=ERROR)
        else:
            dot.config(bg=BORDER,   fg=SUBTEXT)
            lbl.config(fg=SUBTEXT)

    def reset(self):
        for i in range(len(STEPS)):
            self.set(i, STEP_IDLE)


class TooltipMixin:
    def add_tooltip(self, widget, text):
        tip = None
        def show(e):
            nonlocal tip
            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{e.x_root+12}+{e.y_root+6}")
            tk.Label(tip, text=text, bg="#2d3352", fg=TEXT,
                     font=FONT_SM, padx=6, pady=3, relief="flat", bd=1).pack()
        def hide(e):
            nonlocal tip
            if tip:
                tip.destroy()
                tip = None
        widget.bind("<Enter>", show)
        widget.bind("<Leave>", hide)


class GithubUploader(tk.Tk, TooltipMixin):
    def __init__(self):
        super().__init__()
        self.title("CSV / Excel → JSON  ·  GitHub Uploader  |  Hood College Coding Club")
        self.geometry("900x780")
        self.minsize(760, 660)
        self.configure(bg=BG)
        self.resizable(True, True)

        # ── State ──────────────────────────────────────────────────────────
        self.file_path     = tk.StringVar()
        self.sheet_name    = tk.StringVar(value="Sheet1")
        self.github_token  = tk.StringVar()
        self.repo_owner    = tk.StringVar()
        self.repo_name     = tk.StringVar()
        self.remote_path   = tk.StringVar(value="data/output.json")
        self.commit_msg    = tk.StringVar(value="Upload JSON via Hood College Coding Club Uploader")
        self.branch        = tk.StringVar(value="main")
        self.indent_level  = tk.IntVar(value=2)
        self.overwrite_var = tk.BooleanVar(value=True)
        self._is_excel     = False

        self._json_preview  = ""
        self._upload_thread = None

        self._build_ui()
        log_to_file("info", f"Application started. Log file: {LOG_FILE}")

    # ─── UI Construction ──────────────────────────────────────────────────

    def _build_ui(self):
        self._build_header()
        self._build_step_bar()
        main = tk.Frame(self, bg=BG)
        main.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(1, weight=1)
        self._build_file_card(main)
        self._build_github_card(main)
        self._build_preview(main)
        self._build_log(main)
        self._build_footer()

    def _build_header(self):
        hdr = tk.Frame(self, bg=SURFACE, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="CSV / Excel  →  JSON  ·  GitHub Uploader",
                 font=FONT_H1, bg=SURFACE, fg=TEXT).pack(side="left", padx=20)
        # Branding badge
        badge = tk.Frame(hdr, bg="#1e2540", padx=10, pady=4)
        badge.pack(side="right", padx=16)
        tk.Label(badge, text="🎓  Hood Coding Club", font=("Segoe UI", 9, "bold"),
                 bg="#1e2540", fg=ACCENT2).pack()
        tk.Label(badge, text="hood.edu", font=FONT_XS,
                 bg="#1e2540", fg=SUBTEXT).pack()

    def _build_step_bar(self):
        wrapper = tk.Frame(self, bg=SURFACE, pady=8)
        wrapper.pack(fill="x")
        self.step_bar = StepBar(wrapper)
        self.step_bar.pack(padx=20)
        # Mark first two steps immediately (user sets them up before clicking anything)
        self.step_bar.set(0, STEP_IDLE)

    def _card(self, parent, title, row, col, rowspan=1, colspan=1):
        frame = tk.LabelFrame(parent, text=f"  {title}  ",
                              bg=CARD, fg=ACCENT, font=FONT_H2,
                              bd=1, relief="flat",
                              highlightbackground=BORDER,
                              highlightthickness=1)
        frame.grid(row=row, column=col, rowspan=rowspan, columnspan=colspan,
                   sticky="nsew", padx=6, pady=6)
        frame.columnconfigure(1, weight=1)
        return frame

    def _label(self, parent, text, row):
        tk.Label(parent, text=text, bg=CARD, fg=SUBTEXT,
                 font=FONT_SM, anchor="w").grid(
            row=row, column=0, sticky="w", padx=(12, 6), pady=4)

    def _entry(self, parent, var, row, show=None):
        kw = dict(textvariable=var, bg=SURFACE, fg=TEXT,
                  insertbackground=TEXT, relief="flat",
                  font=FONT_UI, bd=0, highlightthickness=1,
                  highlightbackground=BORDER, highlightcolor=ACCENT)
        if show:
            kw["show"] = show
        e = tk.Entry(parent, **kw)
        e.grid(row=row, column=1, sticky="ew", padx=(0, 12), pady=4, ipady=4)
        return e

    def _build_file_card(self, parent):
        f = self._card(parent, "📂  File & Options", row=0, col=0)

        # File picker
        self._label(f, "File", 0)
        row0 = tk.Frame(f, bg=CARD)
        row0.grid(row=0, column=1, sticky="ew", padx=(0, 12), pady=4)
        row0.columnconfigure(0, weight=1)
        self._file_entry = tk.Entry(row0, textvariable=self.file_path,
                 bg=SURFACE, fg=TEXT, relief="flat",
                 insertbackground=TEXT, font=FONT_UI,
                 highlightthickness=1, highlightbackground=BORDER,
                 highlightcolor=ACCENT)
        self._file_entry.grid(row=0, column=0, sticky="ew", ipady=4)
        tk.Button(row0, text="Browse", command=self._pick_file,
                  bg=ACCENT, fg="white", font=FONT_SM,
                  relief="flat", cursor="hand2", padx=10).grid(
            row=0, column=1, padx=(6, 0))
        tk.Label(f, text="Supports .csv  .xlsx  .xls",
                 bg=CARD, fg=SUBTEXT, font=FONT_XS).grid(
            row=1, column=1, sticky="w", padx=(0, 12))

        # Sheet name (only relevant for Excel)
        self._label(f, "Sheet Name", 2)
        self._sheet_entry = self._entry(f, self.sheet_name, 2)
        self.add_tooltip(self._sheet_entry,
            "Excel only: name of the sheet to read.\n"
            "Leave blank to read the first sheet.")
        self._sheet_lbl_widget = f.grid_slaves(row=2, column=0)[0]

        # JSON indent
        self._label(f, "JSON Indent", 3)
        spin_frame = tk.Frame(f, bg=CARD)
        spin_frame.grid(row=3, column=1, sticky="w", padx=(0, 12), pady=4)
        tk.Spinbox(spin_frame, from_=0, to=8, textvariable=self.indent_level,
                   width=5, bg=SURFACE, fg=TEXT, insertbackground=TEXT,
                   relief="flat", font=FONT_UI, buttonbackground=SURFACE).pack(side="left")
        tk.Label(spin_frame, text=" spaces  (0 = minified)",
                 bg=CARD, fg=SUBTEXT, font=FONT_SM).pack(side="left")

        # Remote path
        self._label(f, "Remote Path", 4)
        rp = self._entry(f, self.remote_path, 4)
        self.add_tooltip(rp, "Path inside the repo, e.g.  data/users.json")

        # Update sheet row visibility
        self._update_sheet_row()

    def _update_sheet_row(self):
        color = SUBTEXT if self._is_excel else BORDER
        self._sheet_lbl_widget.config(fg=color)
        self._sheet_entry.config(
            state="normal" if self._is_excel else "disabled",
            fg=TEXT if self._is_excel else SUBTEXT)

    def _build_github_card(self, parent):
        f = self._card(parent, "🐙  GitHub Settings", row=0, col=1)

        self._label(f, "Personal Token", 0)
        e = self._entry(f, self.github_token, 0, show="•")
        self.add_tooltip(e,
            "GitHub → Settings → Developer Settings → Fine-grained tokens\n"
            "Required permission: Contents → Read & Write")

        self._label(f, "Owner / Org", 1)
        owner_entry = self._entry(f, self.repo_owner, 1)
        self.add_tooltip(owner_entry,
            "Just the username/org, e.g. my-org\n"
            "You can also paste a full GitHub URL — it will be parsed automatically.")
        def _on_owner_focusout(e):
            raw = self.repo_owner.get().strip()
            if not raw:
                return
            owner, repo = parse_github_url(raw)
            self.repo_owner.set(owner)
            if repo and not self.repo_name.get().strip():
                self.repo_name.set(repo)
                self._log(f"Auto-filled owner='{owner}', repo='{repo}' from URL", SUBTEXT)
            elif owner != raw:
                self._log(f"Auto-parsed owner → '{owner}'", SUBTEXT)
        owner_entry.bind("<FocusOut>", _on_owner_focusout)

        self._label(f, "Repository", 2)
        self._entry(f, self.repo_name, 2)

        self._label(f, "Branch", 3)
        self._entry(f, self.branch, 3)

        self._label(f, "Commit Message", 4)
        self._entry(f, self.commit_msg, 4)

        self._label(f, "Overwrite?", 5)
        tk.Checkbutton(f, variable=self.overwrite_var,
                       text="Yes, overwrite if file exists",
                       bg=CARD, fg=TEXT, selectcolor=SURFACE,
                       activebackground=CARD, font=FONT_SM,
                       cursor="hand2").grid(
            row=5, column=1, sticky="w", padx=(0, 12), pady=4)

        # ── Action buttons always inside the card ──────────────────────────
        btn_row = tk.Frame(f, bg=CARD)
        btn_row.grid(row=6, column=0, columnspan=2,
                     sticky="ew", padx=12, pady=(10, 8))
        btn_row.columnconfigure(0, weight=1)
        btn_row.columnconfigure(1, weight=1)

        tk.Button(btn_row, text="🔌  Test Connection",
                  command=self._test_connection,
                  bg=CARD, fg=ACCENT2,
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2", pady=7, padx=10,
                  highlightthickness=1,
                  highlightbackground=ACCENT2).grid(
            row=0, column=0, sticky="ew", padx=(0, 4))

        self.upload_btn_card = tk.Button(
            btn_row, text="⬆  Convert & Upload",
            command=self._start_upload,
            bg=ACCENT, fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat", cursor="hand2", pady=7, padx=10)
        self.upload_btn_card.grid(row=0, column=1, sticky="ew", padx=(4, 0))

    def _build_preview(self, parent):
        f = self._card(parent, "🔍  JSON Preview", row=1, col=0)
        f.rowconfigure(0, weight=1)
        f.columnconfigure(0, weight=1)
        self.preview_box = scrolledtext.ScrolledText(
            f, bg=SURFACE, fg=SUCCESS, font=FONT_MONO,
            relief="flat", bd=0, wrap="none", highlightthickness=0)
        self.preview_box.grid(row=0, column=0, columnspan=2,
                              sticky="nsew", padx=8, pady=8)
        self.preview_box.insert("end", "# JSON preview will appear here after loading a file…\n")
        self.preview_box.config(state="disabled")
        parent.rowconfigure(1, weight=1)

    def _build_log(self, parent):
        f = self._card(parent, "📋  Activity Log", row=1, col=1)
        f.rowconfigure(0, weight=1)
        f.columnconfigure(0, weight=1)
        self.log_box = scrolledtext.ScrolledText(
            f, bg=SURFACE, fg=TEXT, font=FONT_MONO,
            relief="flat", bd=0, wrap="word", highlightthickness=0)
        self.log_box.grid(row=0, column=0, columnspan=2,
                          sticky="nsew", padx=8, pady=(8, 4))
        self.log_box.config(state="disabled")

        # Log file location pill
        pill = tk.Frame(f, bg=SURFACE)
        pill.grid(row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))
        tk.Label(pill, text="📁 Log file:", bg=SURFACE, fg=SUBTEXT, font=FONT_XS).pack(side="left")
        tk.Label(pill, text=str(LOG_FILE), bg=SURFACE, fg=ACCENT, font=FONT_XS,
                 cursor="hand2").pack(side="left", padx=(4, 0))

        self._log("Ready. Select a CSV or Excel file to begin.")
        self._log(f"Session log: {LOG_FILE}", SUBTEXT)

    def _build_footer(self):
        bar = tk.Frame(self, bg=SURFACE, pady=8)
        bar.pack(fill="x", side="bottom")

        self.status_lbl = tk.Label(bar, text="●  Idle",
                                   font=FONT_SM, bg=SURFACE, fg=SUBTEXT)
        self.status_lbl.pack(side="left", padx=20)

        self.progress = ttk.Progressbar(bar, mode="indeterminate", length=160)
        self.progress.pack(side="left", padx=8)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TProgressbar", troughcolor=BORDER,
                        background=ACCENT, lightcolor=ACCENT, darkcolor=ACCENT)

        btn_cfg = dict(relief="flat", cursor="hand2",
                       font=("Segoe UI", 9, "bold"), pady=5, padx=14)
        tk.Button(bar, text="🗑  Clear Log", command=self._clear_log,
                  bg=CARD, fg=SUBTEXT, **btn_cfg).pack(side="right", padx=(0, 10))
        tk.Button(bar, text="👁  Preview Only", command=self._preview_only,
                  bg=CARD, fg=TEXT, **btn_cfg).pack(side="right", padx=(0, 6))

        # Branding
        tk.Label(bar, text="Developed by Hood College Coding Club",
                 font=FONT_XS, bg=SURFACE, fg=BORDER).pack(side="right", padx=16)

    # ─── Helpers ─────────────────────────────────────────────────────────

    def _log(self, msg, color=None):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.config(state="normal")
        tag = f"t{id(msg)}{ts}"
        self.log_box.insert("end", f"[{ts}]  {msg}\n", tag)
        if color:
            self.log_box.tag_config(tag, foreground=color)
        self.log_box.see("end")
        self.log_box.config(state="disabled")
        level = "error" if color == ERROR else "info"
        log_to_file(level, msg)

    def _clear_log(self):
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")

    def _set_status(self, text, color=TEXT):
        self.status_lbl.config(text=f"●  {text}", fg=color)

    def _set_upload_btns(self, state):
        self.upload_btn_card.config(state=state)

    # ─── File Picking ─────────────────────────────────────────────────────

    def _pick_file(self):
        path = filedialog.askopenfilename(
            filetypes=[
                ("All Supported", "*.csv *.xlsx *.xls"),
                ("CSV Files",     "*.csv"),
                ("Excel Files",   "*.xlsx *.xls"),
                ("All Files",     "*.*"),
            ])
        if not path:
            return
        self.file_path.set(path)
        ext = Path(path).suffix.lower()
        self._is_excel = ext in (".xlsx", ".xls")
        self._update_sheet_row()
        self._log(f"Selected: {Path(path).name}  ({'Excel' if self._is_excel else 'CSV'})")
        self.step_bar.set(0, STEP_DONE)
        self._load_preview(path)

    # ─── File → JSON Conversion ───────────────────────────────────────────

    @staticmethod
    def _is_empty_row(row: dict) -> bool:
        """Return True if every value in the row is blank/None."""
        return all(
            (v is None or str(v).strip() == "")
            for v in row.values()
        )

    def _file_to_rows(self, path: str) -> list:
        ext = Path(path).suffix.lower()
        if ext in (".xlsx", ".xls"):
            return self._excel_to_rows(path)
        return self._csv_to_rows(path)

    def _csv_to_rows(self, path: str) -> list:
        rows = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if not self._is_empty_row(row):
                    rows.append(dict(row))
        return rows

    def _excel_to_rows(self, path: str) -> list:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel files.\n"
                "Run:  pip install openpyxl")
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_input = self.sheet_name.get().strip()
        if sheet_input and sheet_input in wb.sheetnames:
            ws = wb[sheet_input]
        else:
            ws = wb.active
            if sheet_input:
                self._log(f"⚠  Sheet '{sheet_input}' not found — using '{ws.title}'", WARNING)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for raw in rows[1:]:
            row_dict = {headers[i]: (raw[i] if raw[i] is not None else "") for i in range(len(headers))}
            # Normalise non-serialisable types (dates, etc.)
            for k, v in row_dict.items():
                if hasattr(v, "isoformat"):
                    row_dict[k] = v.isoformat()
                elif v is not None:
                    row_dict[k] = v
            if not self._is_empty_row(row_dict):
                result.append(row_dict)
        return result

    def _rows_to_json(self, rows: list) -> str:
        indent = self.indent_level.get() or None
        return json.dumps(rows, indent=indent, ensure_ascii=False)

    def _load_preview(self, path: str):
        try:
            self._log("Parsing file…")
            rows = self._file_to_rows(path)
            json_str = self._rows_to_json(rows)
            self._json_preview = json_str
            skipped = "(empty rows automatically excluded)"
            self._log(
                f"✅  Parsed {len(rows):,} rows → {len(json_str):,} chars of JSON  {skipped}",
                SUCCESS)
            preview = json_str[:4000]
            if len(json_str) > 4000:
                preview += "\n\n// … (truncated for preview)"
            self.preview_box.config(state="normal")
            self.preview_box.delete("1.0", "end")
            self.preview_box.insert("end", preview)
            self.preview_box.config(state="disabled")
        except Exception as e:
            self._log(f"❌  Parse error: {e}", ERROR)
            self.step_bar.set(0, STEP_ERROR)

    def _preview_only(self):
        path = self.file_path.get().strip()
        if not path:
            messagebox.showwarning("No file", "Please select a CSV or Excel file first.")
            return
        self._load_preview(path)

    # ─── Validation ───────────────────────────────────────────────────────

    def _validate(self):
        errs = []
        if not self.file_path.get().strip():
            errs.append("File path is required.")
        if not self.github_token.get().strip():
            errs.append("GitHub personal access token is required.")
        if not self.repo_owner.get().strip():
            errs.append("Repository owner/org is required.")
        if not self.repo_name.get().strip():
            errs.append("Repository name is required.")
        if not self.remote_path.get().strip():
            errs.append("Remote file path is required.")
        return errs

    # ─── Test Connection ─────────────────────────────────────────────────

    def _test_connection(self):
        token = self.github_token.get().strip()
        if not token:
            messagebox.showwarning("Missing Token",
                "Enter your GitHub Personal Access Token first.")
            return
        raw_owner = self.repo_owner.get().strip()
        raw_repo  = self.repo_name.get().strip()
        owner, maybe_repo = parse_github_url(raw_owner)
        repo = raw_repo or maybe_repo
        self.step_bar.set(2, STEP_ACTIVE)
        self._log("─── Testing GitHub connection… ───", SUBTEXT)
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28",
        }

        def _run():
            try:
                r = requests.get("https://api.github.com/user", headers=headers, timeout=10)
                if r.status_code == 401:
                    self.after(0, lambda: self._log("❌  Token invalid — 401 Unauthorized", ERROR))
                    self.after(0, lambda: self.step_bar.set(2, STEP_ERROR))
                    return
                if r.status_code == 200:
                    login = r.json().get("login", "?")
                    self.after(0, lambda: self._log(
                        f"✅  Token valid — authenticated as: {login}", SUCCESS))
                else:
                    self.after(0, lambda: self._log(
                        f"⚠  /user returned {r.status_code}", ERROR))
                    self.after(0, lambda: self.step_bar.set(2, STEP_ERROR))
                    return

                if not owner or not repo:
                    self.after(0, lambda: self._log(
                        "⚠  Fill in Owner and Repository to test repo access.", WARNING))
                    self.after(0, lambda: self.step_bar.set(2, STEP_ERROR))
                    return

                r2 = requests.get(
                    f"https://api.github.com/repos/{owner}/{repo}",
                    headers=headers, timeout=10)
                if r2.status_code == 200:
                    d = r2.json()
                    vis = "private" if d.get("private") else "public"
                    db  = d.get("default_branch", "?")
                    self.after(0, lambda: self._log(
                        f"✅  Repo found: {owner}/{repo}  [{vis}]  default_branch={db}", SUCCESS))
                    self.after(0, lambda: self.step_bar.set(2, STEP_DONE))
                    self.after(0, lambda: self.step_bar.set(1, STEP_DONE))
                elif r2.status_code == 404:
                    self.after(0, lambda: self._log(
                        f"❌  Repo not found: {owner}/{repo}\n"
                        "    Check spelling, or token may lack Contents permission.", ERROR))
                    self.after(0, lambda: self.step_bar.set(2, STEP_ERROR))
                elif r2.status_code == 403:
                    self.after(0, lambda: self._log(
                        "❌  403 Forbidden — token lacks Contents permission.", ERROR))
                    self.after(0, lambda: self.step_bar.set(2, STEP_ERROR))
                else:
                    self.after(0, lambda: self._log(
                        f"⚠  Repo check returned {r2.status_code}", ERROR))
                    self.after(0, lambda: self.step_bar.set(2, STEP_ERROR))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: self._log(f"❌  Connection error: {err}", ERROR))
                self.after(0, lambda: self.step_bar.set(2, STEP_ERROR))

        threading.Thread(target=_run, daemon=True).start()

    # ─── Upload ───────────────────────────────────────────────────────────

    def _start_upload(self):
        errs = self._validate()
        if errs:
            messagebox.showerror("Validation Error", "\n".join(errs))
            return
        if self._upload_thread and self._upload_thread.is_alive():
            return
        json_bytes = len(self._json_preview.encode()) if self._json_preview else 0
        if json_bytes > 900_000:
            if not messagebox.askyesno(
                "Large File Warning",
                f"The JSON is ~{json_bytes // 1024} KB.\n"
                "GitHub's Contents API has a 1 MB limit.\n\n"
                "Try JSON Indent = 0 (minified) to reduce size.\n\nProceed anyway?"):
                return
        self.step_bar.reset()
        self.step_bar.set(0, STEP_DONE)
        self._set_upload_btns("disabled")
        self.progress.start(12)
        self._set_status("Uploading…", ACCENT)
        self._log("═══════════════════════════════════", SUBTEXT)
        self._log("Starting upload session…", SUBTEXT)
        log_to_file("info", "Upload session started")
        self._upload_thread = threading.Thread(target=self._upload_worker, daemon=True)
        self._upload_thread.start()

    def _upload_worker(self):
        try:
            path  = self.file_path.get().strip()
            token = self.github_token.get().strip()
            branch = self.branch.get().strip() or "main"
            fpath  = self.remote_path.get().strip()
            msg    = self.commit_msg.get().strip() or "Upload JSON"

            raw_owner = self.repo_owner.get().strip()
            raw_repo  = self.repo_name.get().strip()
            owner, maybe_repo = parse_github_url(raw_owner)
            if maybe_repo and not raw_repo:
                raw_repo = maybe_repo
            repo = raw_repo

            if not owner or not repo:
                raise ValueError(
                    f"Could not determine owner/repo.\n"
                    f"  Owner: '{raw_owner}'\n  Repo: '{raw_repo}'"
                )

            # Step 3 — Convert
            self.after(0, lambda: self.step_bar.set(3, STEP_ACTIVE))
            self.after(0, lambda: self._log(
                f"[Step 3/5]  Converting {Path(path).name}…"))
            rows = self._file_to_rows(path)
            json_str = self._rows_to_json(rows)
            self._json_preview = json_str
            self.after(0, self._refresh_preview)
            self.after(0, lambda: self._log(
                f"  ✅  Converted {len(rows):,} rows → {len(json_str):,} chars", SUCCESS))
            self.after(0, lambda: self.step_bar.set(3, STEP_DONE))

            headers = {
                "Authorization": f"Bearer {token}",
                "Accept": "application/vnd.github+json",
                "X-GitHub-Api-Version": "2022-11-28",
            }
            api = f"https://api.github.com/repos/{owner}/{repo}/contents/{fpath}"

            # Step 4 — Check remote
            self.after(0, lambda: self.step_bar.set(4, STEP_ACTIVE))
            self.after(0, lambda: self._log(
                f"[Step 4/5]  Checking remote: {owner}/{repo}/{fpath}"))
            sha = None
            r = requests.get(api, headers=headers, params={"ref": branch}, timeout=15)
            if r.status_code == 200:
                if not self.overwrite_var.get():
                    self.after(0, lambda: messagebox.showerror(
                        "File Exists", "Remote file already exists.\nEnable 'Overwrite'."))
                    self.after(0, lambda: self.step_bar.set(4, STEP_ERROR))
                    return
                sha = r.json().get("sha")
                self.after(0, lambda: self._log(
                    "  Remote file exists — will overwrite.", SUBTEXT))
            elif r.status_code == 404:
                self.after(0, lambda: self._log(
                    "  No existing remote file — will create new.", SUBTEXT))
            elif r.status_code == 401:
                raise PermissionError("401 Unauthorized — check your token.")
            elif r.status_code == 403:
                raise PermissionError("403 Forbidden — token lacks Contents: Write permission.")
            else:
                raise RuntimeError(
                    f"GitHub API error {r.status_code}\n{r.text[:300]}")

            # Step 5 — Upload
            self.after(0, lambda: self._log(
                f"[Step 5/5]  Uploading to GitHub… ({len(json_str) // 1024 + 1} KB)"))
            encoded = base64.b64encode(json_str.encode()).decode()
            payload = {"message": msg, "content": encoded, "branch": branch}
            if sha:
                payload["sha"] = sha

            r2 = requests.put(api, headers=headers, json=payload, timeout=30)

            if r2.status_code in (200, 201):
                info     = r2.json()
                html_url = info.get("content", {}).get("html_url", "")
                commit_url = info.get("commit", {}).get("html_url", "")
                self.after(0, lambda: self._log(
                    f"  ✅  File:   {html_url}", SUCCESS))
                self.after(0, lambda: self._log(
                    f"  ✅  Commit: {commit_url}", SUCCESS))
                self.after(0, lambda: self.step_bar.set(4, STEP_DONE))
                self.after(0, lambda: self.step_bar.set(5, STEP_DONE))
                self.after(0, lambda: self._set_status("Upload complete!", SUCCESS))
                self.after(0, lambda: messagebox.showinfo(
                    "Done!",
                    f"JSON uploaded successfully!\n\n"
                    f"File:   {html_url}\n"
                    f"Commit: {commit_url}"))
                log_to_file("info",
                    f"SUCCESS — {owner}/{repo}/{fpath}  rows={len(rows)}  bytes={len(json_str)}")
            elif r2.status_code == 401:
                raise PermissionError("401 Unauthorized on upload.")
            elif r2.status_code == 422:
                raise RuntimeError(f"422 Unprocessable — SHA conflict or bad branch.\n{r2.text[:300]}")
            else:
                raise RuntimeError(f"Upload failed {r2.status_code}\n{r2.text[:300]}")

        except Exception as e:
            err = str(e)
            self.after(0, lambda: self._log(f"❌  Error: {err}", ERROR))
            self.after(0, lambda: self._set_status("Error", ERROR))
            self.after(0, lambda: messagebox.showerror("Upload Failed", err))
            log_to_file("error", f"Upload failed: {err}")
        finally:
            self.after(0, self._upload_done)

    def _upload_done(self):
        self.progress.stop()
        self._set_upload_btns("normal")
        if "complete" not in self.status_lbl.cget("text"):
            self._set_status("Idle")

    def _refresh_preview(self):
        preview = self._json_preview[:4000]
        if len(self._json_preview) > 4000:
            preview += "\n\n// … (truncated)"
        self.preview_box.config(state="normal")
        self.preview_box.delete("1.0", "end")
        self.preview_box.insert("end", preview)
        self.preview_box.config(state="disabled")


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    missing = []
    try:
        import requests  # noqa: F401
    except ImportError:
        missing.append("requests")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")

    if missing:
        import subprocess, sys
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install"] + missing)

    app = GithubUploader()
    app.mainloop()